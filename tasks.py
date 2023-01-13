import json

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from api_client import YandexWeatherAPI
from models import (
    CityModel, DayWeatherConditionsModel,
    CombinedWeatherConditionsModel, FinalOutputModel
)
from utils import RESULT_FILE_NAME, START_TIME, END_TIME, CONDITIONS, logger, COLUMNS_NAME


class DataFetchingTask:
    """
    Получение данных через API
    """
    yw_api: YandexWeatherAPI = YandexWeatherAPI()

    def make_request(self, city_name: str) -> CityModel:
        """Получение данных по API"""
        logger.info("Making request to Yandex Weather API for city: %s",
                    city_name)
        city_data = self.yw_api.get_forecasting(city_name)
        logger.debug("API response: %s", city_data)
        return CityModel(
            city=city_name,
            forecasts=city_data
        )


class DataCalculationTask:
    """
    Вычисление погодных параметров
    """

    @staticmethod
    def __additional_calculating(temp_data: list) -> float:
        """ Вычисление средних показателей с округлением."""

        if not temp_data:
            return 0
        return round(sum(temp_data) / len(temp_data), 2)

    def _calculating_data(self, city_data: CityModel) -> \
            CombinedWeatherConditionsModel:
        """ Вычисление средних показателей"""

        logger.info("Calculating statistics for city: %s", city_data.city)
        weather_data = []
        for day in city_data.forecasts.forecasts:
            result_temp_for_day = []
            count_clear_cond = 0

            for hour in day.hours:
                if START_TIME < hour.hour < END_TIME:
                    result_temp_for_day.append(hour.temp)
                    count_clear_cond += \
                        1 if hour.condition in CONDITIONS else 0
            logger.debug("Creating and adding a DayWeatherConditionsModel to "
                         "a list for city: %s", city_data.city)
            weather_data.append(
                DayWeatherConditionsModel(
                    date=day.date,
                    daily_avg_temp=self.__additional_calculating(
                        result_temp_for_day),
                    clear_weather_cond=count_clear_cond
                ))

        return CombinedWeatherConditionsModel(
            city=city_data.city,
            data=weather_data,
        )

    def _calc_general_indicators_for_day(self,
                                         data:
                                         CombinedWeatherConditionsModel) -> \
            FinalOutputModel:
        """Добавление финальных показателей"""

        list_avg_temp = []
        total_avg_clear_weather_cond = 0
        for element in data.data:
            if element.daily_avg_temp != 0.0:
                list_avg_temp.append(element.daily_avg_temp)
                total_avg_clear_weather_cond += element.clear_weather_cond

        logger.debug("Calculation of averages for city: %s", data.city)
        total_avg_temperature = self.__additional_calculating(list_avg_temp)

        logger.debug("Creating and returning a FinalOutputModel to "
                     "for city: %s", data.city)
        return FinalOutputModel(
            city=data.city,
            data=data.data,
            total_avg_temp=total_avg_temperature,
            total_clear_weather_cond=total_avg_clear_weather_cond,
            rating=0
        )

    def general_calculation(self,
                            city_data: CityModel) -> FinalOutputModel:
        """Вычисление результатов для обработки вне класса"""
        logger.info("Getting the FinalOutputModel for city: %s",
                    city_data.city)

        calculating = self._calculating_data(city_data)
        result = self._calc_general_indicators_for_day(calculating)
        logger.debug("Returning the FinalOutputModel for city: %s",
                     city_data.city)
        return result

    @staticmethod
    def adding_rating(data: list[FinalOutputModel]) -> list[dict]:
        """Добавление рейтинга города"""

        total_rating = [
            (one_model.city,
             one_model.total_avg_temp,
             one_model.total_clear_weather_cond) for one_model in data
        ]

        sorted_rating = sorted(total_rating, key=lambda x: (-x[1], -x[2]))

        dict_rating = {
            a[0]: i + 1 for i, a in enumerate(sorted_rating)}
        logger.debug("Adding a rating to each city: %s",
                     dict_rating)

        for element in data:
            element.rating = dict_rating[element.city]

        result = [element.dict() for element in data]
        logger.debug("Converting models to a dictionary")

        return result


class DataAggregationTask:
    """
    Объединение вычисленных данных
    """

    @staticmethod
    def save_results_as_json(to_save: list[dict]) -> None:
        """Сохранение аналитических данных в формате Json"""

        logger.debug("Saving results to a file")
        with open(RESULT_FILE_NAME, 'w', encoding='utf-8') as file:
            json.dump(to_save, file, indent=2)

    @staticmethod
    def prepare_table_xlsx(data: list[dict]) -> None:
        """Создание таблицы и заполнение первой строки"""

        wb = Workbook()
        sheet = wb.active
        data_first_string = ['Страна/день', 'Показатель', 'Среднее', 'Рейтинг']
        elm_from_data = data[0].get('data')
        amount_city = len(data)
        list_of_days = [i.get('date') for i in elm_from_data]
        for idx in range(len(list_of_days)):
            data_first_string.insert(2 + idx, list_of_days[idx])
        for letter in COLUMNS_NAME:
            sheet.column_dimensions[letter].width = 15
        for row in range((amount_city + 1) * 2):
            sheet.row_dimensions[row].height = 20
        sheet.append(data_first_string)
        wb.save('test_tmp_file.xlsx')

    @staticmethod
    def preparing_data_for_insertion(element: dict) -> (list, list):
        """Подготовка данных для вставки в таблицу"""

        row_1 = [
            element.get('city'),
            'Температура, среднее',
            *(element.get('data')[i].get('daily_avg_temp') for i in
              range(len(element.get('data')))),
            element.get('total_avg_temp'),
            element.get('rating')
        ]
        row_2 = [
            '',
            'Без осадков, часов',
            *(element.get('data')[i].get('clear_weather_cond') for i in
              range(len(element.get('data')))),
            element.get('total_clear_weather_cond'),
            ''
        ]

        return row_1, row_2

    @staticmethod
    def filling_in_table(element: list | tuple) -> None:
        """Вставка данных в таблицу"""

        wb = openpyxl.load_workbook('test_tmp_file.xlsx')
        sheet = wb.active
        sheet.append(element[0])
        sheet.append(element[1])
        wb.save('test_tmp_file.xlsx')

    @staticmethod
    def adding_boarder():
        """Добавление разметки между столбцами и колонками"""

        wb = openpyxl.load_workbook('test_tmp_file.xlsx')
        sheet = wb.active
        thins = Side(border_style="thin", color="000000")
        for row in range(1, 33):
            for col in range(1, 10):
                if row == 32:
                    sheet.cell(row=row, column=col).border = Border(top=thins)
                else:
                    sheet.cell(
                        row=row, column=col).border = Border(
                        left=thins,
                        right=thins,
                        top=thins,
                        end=thins)
        wb.save('test_tmp_file.xlsx')


class DataAnalyzingTask:
    """
     Финальный анализ и получение результата
    """

    @staticmethod
    def get_result(data: list[dict]) -> str:
        """Получение результата"""

        logger.info("Finding the best temperature and number of sunny days")
        result = []
        max_temp = -9999
        sunniest_weather = 0
        for city_data in data:
            if city_data['rating'] == 1:
                max_temp = city_data['total_avg_temp']
                sunniest_weather = city_data['total_clear_weather_cond']
                break

        logger.info("Adding eligible cities to the final list")
        for city_data in data:
            if city_data['total_avg_temp'] == max_temp or \
                    city_data['total_clear_weather_cond'] == sunniest_weather:
                result.append(
                    {
                        'city':
                            city_data['city'],
                        'total_avg_temp':
                            city_data['total_avg_temp'],
                        'total_clear_weather_cond':
                            city_data['total_clear_weather_cond']
                    }
                )
        answer = 'Самые лучшие погодные условия получились в городе'
        result_str = ''
        for city in result:
            result_str += f' {city["city"]}, средняя температура: ' \
                          f'{city["total_avg_temp"]}, а количество ' \
                          f'времени без осадков ' \
                          f'{city["total_clear_weather_cond"]} часов.\n'
        return answer + result_str
